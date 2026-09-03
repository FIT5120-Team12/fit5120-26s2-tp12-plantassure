"""
analysis.py — Iteration 1 Data Analysis

A standalone script, separate from pipeline.py, that reproduces and
visualises the four analyses described in the Data Management Plan,
Section 8:

  8.1 Cross-Dataset Overlap Verification
  8.2 Subspecies Collision Analysis
  8.3 Risk Category Verification
  8.4 Assessment Coverage Analysis

It imports the real matching/rule functions from pipeline.py rather
than re-implementing them, so these findings can never silently drift
out of sync with the actual pipeline behaviour.

Usage:
    python3 analysis.py

Outputs (written to analysis_output/):
    8.1_overlap.png
    8.2_subspecies_collisions.png
    8.3_risk_categories.png
    8.4_assessment_coverage.png
    8.4_recommendation_distribution.png
    findings.txt   (plain-text summary of every number used above)
"""
import os
import openpyxl
import shapefile
import pandas as pd
import matplotlib.pyplot as plt

from pipeline import (
    normalize_name_species,
    normalize_name_specific,
    run_pipeline,
)

OUT_DIR = "analysis_output"

# Brand palette, matching the Data Management Plan / eportfolio
TEAL = "#1B3B33"
AMBER = "#E8793C"
LIGHT_TEAL = "#3E97A6"
GREY = "#B4B2A9"
RED = "#A32D2D"

VICFLORA_PATH = "data/vicflora_monash_2026.csv"
ADVISORY_PATH = "data/Advisory-list-of-environmental-weeds-in-Victoria_2022.xlsx"
VBA_PATH = "data/Order_ELFIES/mga2020_55/esrishape/lga_polygon/MONASH-0/FLORAFAUNA1/VBA_FLORA25.shp"
GBIF_PATH = "data/iNaturalist.csv"


def style_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#333333")



# 8.1 Cross-Dataset Overlap Verification
def analysis_overlap():
    print("8.1: Cross-dataset overlap verification ...")

    vba_keys = set()
    sf = shapefile.Reader(VBA_PATH)
    field_names = [f[0] for f in sf.fields[1:]]
    idx_name = field_names.index("SCI_NAME")
    for rec in sf.iterRecords():
        name = rec[idx_name]
        if name and str(name).strip():
            key = normalize_name_species(name)
            if key:
                vba_keys.add(key)

    adv_keys = set()
    wb = openpyxl.load_workbook(ADVISORY_PATH, read_only=True, data_only=True)
    ws = wb["Advisory list 2022"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx_name = header.index("Scientific name")
    for row in it:
        name = row[idx_name]
        if name:
            key = normalize_name_species(name)
            if key:
                adv_keys.add(key)

    overlap = vba_keys & adv_keys

    fig, ax = plt.subplots(figsize=(7, 4.2))
    labels = ["VBA\n(Monash extract)", "Advisory List\n(statewide)", "Overlap"]
    values = [len(vba_keys), len(adv_keys), len(overlap)]
    colors = [LIGHT_TEAL, AMBER, TEAL]
    bars = ax.bar(labels, values, color=colors, width=0.55)
    for b, v in zip(bars, values):
        ax.text(b.get_x() + b.get_width() / 2, v + 15, str(v), ha="center", fontsize=11, color="#222222")
    ax.set_ylabel("Unique species")
    ax.set_title("8.1  VBA \u2194 Advisory List overlap (species-level match)", fontsize=12, color=TEAL, weight="bold")
    style_axes(ax)
    plt.tight_layout()
    plt.savefig(f"{OUT_DIR}/7.1_overlap.png", dpi=150)
    plt.close()

    return len(vba_keys), len(adv_keys), len(overlap)



# 8.2 Subspecies Collision Analysis
def analysis_subspecies_collisions():
    print("8.2: Subspecies collision analysis ...")

    wb = openpyxl.load_workbook(ADVISORY_PATH, read_only=True, data_only=True)
    ws = wb["Advisory list 2022"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx_name = header.index("Scientific name")
    idx_risk = header.index("Risk Rating")

    from collections import defaultdict
    by_species_key = defaultdict(set)
    for row in it:
        name = row[idx_name]
        if not name:
            continue
        species_key = normalize_name_species(name)
        if not species_key:
            continue
        risk = row[idx_risk]
        by_species_key[species_key].add((str(name).strip(), str(risk).strip() if risk else None))

    collisions = {}
    for key, entries in by_species_key.items():
        if len(entries) > 1:
            ratings = set(r for _, r in entries)
            if len(ratings) > 1:
                collisions[key] = entries

    fig, ax = plt.subplots(figsize=(7, 4.2))
    labels = ["Species with a single,\nunambiguous rating", "Species affected by the\nsubspecies collision"]
    total_species = len(by_species_key)
    affected = len(collisions)
    values = [total_species - affected, affected]
    colors = [LIGHT_TEAL, RED]
    bars = ax.bar(labels, values, color=colors, width=0.5)
    for b, v in zip(bars, values):
        ax.text(b.get_x() + b.get_width() / 2, v + 10, str(v), ha="center", fontsize=11, color="#222222")
    ax.set_ylabel("Species (Advisory List, species-level key)")
    ax.set_title("8.2  Impact of species-only matching on the Advisory List", fontsize=12, color=TEAL, weight="bold")
    style_axes(ax)
    plt.tight_layout()
    plt.savefig(f"{OUT_DIR}/7.2_subspecies_collisions.png", dpi=150)
    plt.close()

    return collisions



# 8.3 Risk Category Verification
def analysis_risk_categories(result):
    print("7.3: Risk category verification ...")

    from collections import Counter
    verdict_counts = Counter(r["recommendation"] for r in result)

    order = ["Reconsider Planting", "Use Caution", "Lower Concern", "Not Assessed"]
    values = [verdict_counts.get(v, 0) for v in order]
    colors = [RED, AMBER, LIGHT_TEAL, GREY]

    fig, ax = plt.subplots(figsize=(7, 4.2))
    bars = ax.bar(order, values, color=colors, width=0.55)
    for b, v in zip(bars, values):
        ax.text(b.get_x() + b.get_width() / 2, v + 8, str(v), ha="center", fontsize=11, color="#222222")
    ax.set_ylabel("Species")
    ax.set_title("8.3  Final verdict distribution, all 6 Risk Rating categories mapped", fontsize=12, color=TEAL, weight="bold")
    style_axes(ax)
    plt.tight_layout()
    plt.savefig(f"{OUT_DIR}/7.3_risk_categories.png", dpi=150)
    plt.close()

    return verdict_counts



# 8.4 Assessment Coverage Analysis
def analysis_coverage(result):
    print("8.4: Assessment coverage analysis ...")

    total = len(result)
    not_assessed = sum(1 for r in result if r["recommendation"] == "Not Assessed")
    assessed = total - not_assessed

    fig, ax = plt.subplots(figsize=(5.5, 5.5))
    sizes = [assessed, not_assessed]
    labels = [f"Assessed\n{assessed} species ({assessed/total:.0%})", f"Not Assessed\n{not_assessed} species ({not_assessed/total:.0%})"]
    colors = [LIGHT_TEAL, GREY]
    ax.pie(sizes, labels=labels, colors=colors, autopct=None, startangle=90,
           wedgeprops={"edgecolor": "white", "linewidth": 2},
           textprops={"fontsize": 11, "color": "#222222"})
    ax.set_title("8.4  Advisory List assessment coverage\n(Iteration 1, City of Monash extract)", fontsize=12, color=TEAL, weight="bold")
    plt.tight_layout()
    plt.savefig(f"{OUT_DIR}/7.4_assessment_coverage.png", dpi=150)
    plt.close()

    return assessed, not_assessed, total



# 8.5 Foresight: Early-Warning Candidates Among Unassessed Species
def analysis_foresight_candidates(merged_df):
    print("8.5: Foresight - early-warning candidates among unassessed species ...")

    unassessed = merged_df[
        merged_df["risk_rating"].isna()
        | (merged_df["risk_rating"] == "Not Assessed / No exact match")
    ].copy()
    total_unassessed = len(unassessed)

    # Native species are excluded: only introduced/naturalised species can
    # plausibly become future environmental weeds. Including natives here
    # would just surface popular, commonly-photographed wildflowers (an
    # artifact of citizen-science popularity), not a meaningful signal.
    introduced_unassessed = unassessed[
        unassessed["establishment_means"].str.lower().isin(["introduced", "naturalised", "naturalized"])
    ].copy()
    n_introduced = len(introduced_unassessed)

    introduced_unassessed["inat_most_recent_date"] = pd.to_datetime(
        introduced_unassessed["inat_most_recent_date"], errors="coerce"
    )
    has_inat = introduced_unassessed[introduced_unassessed["inat_record_count"] > 0]
    n_with_inat = len(has_inat)

    recent_cutoff = pd.Timestamp("2024-01-01")
    recent = has_inat[has_inat["inat_most_recent_date"] >= recent_cutoff]
    n_recent = len(recent)

    top10 = has_inat.sort_values("inat_record_count", ascending=False).head(10)

    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    names = [n.split()[0][:1] + ". " + " ".join(n.split()[1:]) for n in top10["scientific_name"]]
    values = top10["inat_record_count"].tolist()
    bars = ax.barh(names[::-1], values[::-1], color=TEAL)
    for b, v in zip(bars, values[::-1]):
        ax.text(v + max(values) * 0.01, b.get_y() + b.get_height() / 2, str(int(v)), va="center", fontsize=10, color="#222222")
    ax.set_xlabel("iNaturalist record count")
    ax.set_title("8.5  Unassessed, introduced species by local observation volume", fontsize=12, color=TEAL, weight="bold")
    style_axes(ax)
    plt.tight_layout()
    plt.savefig(f"{OUT_DIR}/8.5_foresight_candidates.png", dpi=150)
    plt.close()

    return {
        "total_unassessed": total_unassessed,
        "n_introduced": n_introduced,
        "n_with_inat": n_with_inat,
        "n_recent": n_recent,
        "top10": top10[["scientific_name", "vernacular_name", "inat_record_count", "inat_most_recent_date"]],
    }



# Main
def main():
    try:
        os.makedirs(OUT_DIR, exist_ok=True)
    except OSError as e:
        print(f"Could not create '{OUT_DIR}/' here: {e}")
        print(f"Current directory: {os.getcwd()}")
        print("This usually means analysis.py was run from the wrong folder.")
        print("cd into the folder containing pipeline.py and the data/ folder, then rerun:")
        print("  cd /path/to/FIT5120_Project-2/pipeline")
        print("  python3 analysis.py")
        return

    print("Running the real pipeline once, so every chart below uses")
    print("exactly the same data pipeline.py produces for the app.\n")
    result, merged_df = run_pipeline(
        vicflora_path=VICFLORA_PATH,
        advisory_path=ADVISORY_PATH,
        vba_shp_path=VBA_PATH,
        gbif_path=GBIF_PATH,
    )
    print()

    vba_n, adv_n, overlap_n = analysis_overlap()
    collisions = analysis_subspecies_collisions()
    verdict_counts = analysis_risk_categories(result)
    assessed, not_assessed, total = analysis_coverage(result)
    foresight = analysis_foresight_candidates(merged_df)

    lines = []
    lines.append("Iteration 1 Data Analysis — Findings\n")
    lines.append("7.1 Cross-Dataset Overlap Verification")
    lines.append(f"  VBA (Monash extract) unique species: {vba_n}")
    lines.append(f"  Advisory List unique species:        {adv_n}")
    lines.append(f"  Overlap (species-level key match):   {overlap_n}\n")

    lines.append("7.2 Subspecies Collision Analysis")
    lines.append(f"  Species affected by species-only matching: {len(collisions)}")
    for key, entries in sorted(collisions.items()):
        lines.append(f"    {key}:")
        for name, risk in sorted(entries):
            lines.append(f"      {name}: {risk}")
    lines.append("")

    lines.append("7.3 Risk Category / Verdict Distribution")
    for v, c in verdict_counts.items():
        lines.append(f"  {v}: {c}")
    lines.append("")

    lines.append("7.4 Assessment Coverage")
    lines.append(f"  Total species (Iteration 1 output): {total}")
    lines.append(f"  Assessed:                            {assessed} ({assessed/total:.1%})")
    lines.append(f"  Not Assessed:                        {not_assessed} ({not_assessed/total:.1%})")
    lines.append("")

    lines.append("8.5 Foresight: Early-Warning Candidates Among Unassessed Species")
    lines.append(f"  Total unassessed species:                     {foresight['total_unassessed']}")
    lines.append(f"  Of those, introduced/naturalised:             {foresight['n_introduced']}")
    lines.append(f"  Of those, with >=1 iNaturalist record:        {foresight['n_with_inat']}")
    lines.append(f"  Of those, with a recent (2024+) observation:  {foresight['n_recent']}")
    lines.append("  Top 10 by iNaturalist record count:")
    for _, r in foresight["top10"].iterrows():
        cn = r["vernacular_name"] if pd.notna(r["vernacular_name"]) else "(no common name)"
        lines.append(f"    {r['scientific_name']} ({cn}): {int(r['inat_record_count'])} records, most recent {r['inat_most_recent_date']}")

    with open(f"{OUT_DIR}/findings.txt", "w") as f:
        f.write("\n".join(lines))

    print("\n".join(lines))
    print(f"\nCharts and findings.txt written to {OUT_DIR}/")


if __name__ == "__main__":
    main()
